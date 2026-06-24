import io
import re
from datetime import date, datetime, time
from urllib.parse import urlparse
from zoneinfo import ZoneInfo, ZoneInfoNotFoundError

from openpyxl import load_workbook


INFO_FIELDS = {
    "organization_name",
    "organization_slug",
    "event_name",
    "event_slug",
    "start_date",
    "end_date",
    "theme",
    "venue_name",
    "venue_address",
    "travel_info",
    "local_activities",
    "website_url",
    "timezone",
    "primary_color",
    "secondary_color",
    "accent_color",
    "logo_url",
    "support_contact",
    "welcome_text",
}

SESSION_COLUMNS = [
    "session_number",
    "session_name",
    "long_description",
    "resources_url",
    "presenters",
    "date",
    "start_time",
    "end_time",
    "room",
    "capacity",
    "track",
    "session_type",
    "cech",
    "featured",
    "visible",
]

REQUIRED_SESSION_COLUMNS = {
    "session_name",
    "date",
    "start_time",
    "end_time",
}

SLUG_RE = re.compile(r"^[a-z0-9]+(?:-[a-z0-9]+)*$")
COLOR_RE = re.compile(r"^#[0-9A-Fa-f]{6}$")


class WorkbookValidationError(ValueError):
    def __init__(self, errors):
        super().__init__("; ".join(errors))
        self.errors = errors


def slugify(value):
    value = re.sub(r"[^a-z0-9]+", "-", str(value).lower()).strip("-")
    return value or "event"


def _string(value):
    if value is None:
        return ""
    return str(value).strip()


def _date(value, field, errors):
    if isinstance(value, datetime):
        return value.date().isoformat()
    if isinstance(value, date):
        return value.isoformat()
    text = _string(value)
    for pattern in ("%Y-%m-%d", "%m/%d/%Y", "%m/%d/%y"):
        try:
            return datetime.strptime(text, pattern).date().isoformat()
        except ValueError:
            pass
    errors.append(f"{field} must be a date such as 2026-04-21.")
    return ""


def _time(value, field, errors):
    if isinstance(value, datetime):
        return value.time().replace(second=0, microsecond=0).isoformat(timespec="minutes")
    if isinstance(value, time):
        return value.replace(second=0, microsecond=0).isoformat(timespec="minutes")
    text = _string(value)
    for pattern in ("%H:%M", "%I:%M %p", "%I:%M%p"):
        try:
            return datetime.strptime(text, pattern).time().isoformat(timespec="minutes")
        except ValueError:
            pass
    errors.append(f"{field} must be a time such as 9:00 AM.")
    return ""


def _boolean(value, default=True):
    if value is None or _string(value) == "":
        return default
    if isinstance(value, bool):
        return value
    return _string(value).lower() in {"yes", "y", "true", "1"}


def _url(value, field, errors):
    text = _string(value)
    if not text:
        return ""
    parsed = urlparse(text)
    if parsed.scheme not in {"http", "https"} or not parsed.netloc:
        errors.append(f"{field} must be a complete http or https URL.")
    return text


def parse_workbook(workbook_bytes):
    errors = []
    try:
        workbook = load_workbook(io.BytesIO(workbook_bytes), data_only=True)
    except Exception as exc:
        raise WorkbookValidationError([f"Could not read the Excel file: {exc}"]) from exc

    required_sheets = {"Conference Info", "Sessions"}
    missing_sheets = required_sheets.difference(workbook.sheetnames)
    if missing_sheets:
        raise WorkbookValidationError(
            [f"Missing worksheet: {sheet}" for sheet in sorted(missing_sheets)]
        )

    info_sheet = workbook["Conference Info"]
    info = {}
    for row in info_sheet.iter_rows(min_row=2, values_only=True):
        key = _string(row[0] if row else None).lower()
        if key:
            info[key] = row[1] if len(row) > 1 else None

    for field in sorted(INFO_FIELDS):
        info.setdefault(field, "")

    if not _string(info["organization_name"]):
        errors.append("Conference Info: organization_name is required.")
    if not _string(info["event_name"]):
        errors.append("Conference Info: event_name is required.")

    info["organization_name"] = _string(info["organization_name"])
    info["event_name"] = _string(info["event_name"])
    info["organization_slug"] = _string(info["organization_slug"]) or slugify(
        info["organization_name"]
    )
    info["event_slug"] = _string(info["event_slug"]) or slugify(info["event_name"])

    for field in ("organization_slug", "event_slug"):
        if not SLUG_RE.fullmatch(info[field]):
            errors.append(
                f"Conference Info: {field} must use lowercase letters, numbers, and hyphens."
            )

    info["start_date"] = _date(info["start_date"], "Conference Info: start_date", errors)
    info["end_date"] = _date(info["end_date"], "Conference Info: end_date", errors)
    for field in (
        "theme",
        "venue_name",
        "venue_address",
        "travel_info",
        "local_activities",
        "support_contact",
        "welcome_text",
    ):
        info[field] = _string(info[field])

    for field in ("website_url", "logo_url"):
        info[field] = _url(info[field], f"Conference Info: {field}", errors)

    info["timezone"] = _string(info["timezone"]) or "America/Los_Angeles"
    try:
        ZoneInfo(info["timezone"])
    except ZoneInfoNotFoundError:
        errors.append(f"Conference Info: unknown timezone {info['timezone']}.")

    color_defaults = {
        "primary_color": "#173B57",
        "secondary_color": "#E7F1F4",
        "accent_color": "#E39A3B",
    }
    for field, default in color_defaults.items():
        info[field] = _string(info[field]) or default
        if not COLOR_RE.fullmatch(info[field]):
            errors.append(f"Conference Info: {field} must look like #173B57.")

    session_sheet = workbook["Sessions"]
    headers = [
        _string(cell.value).lower()
        for cell in next(session_sheet.iter_rows(min_row=1, max_row=1))
    ]
    missing_columns = REQUIRED_SESSION_COLUMNS.difference(headers)
    if missing_columns:
        errors.extend(
            f"Sessions: missing required column {column}."
            for column in sorted(missing_columns)
        )
        raise WorkbookValidationError(errors)

    sessions = []
    for row_number, row in enumerate(
        session_sheet.iter_rows(min_row=2, values_only=True), start=2
    ):
        values = dict(zip(headers, row))
        if not any(_string(value) for value in row):
            continue

        prefix = f"Sessions row {row_number}"
        session_name = _string(values.get("session_name"))
        if not session_name:
            errors.append(f"{prefix}: session_name is required.")

        session_date = _date(values.get("date"), f"{prefix}: date", errors)
        start_time = _time(values.get("start_time"), f"{prefix}: start_time", errors)
        end_time = _time(values.get("end_time"), f"{prefix}: end_time", errors)

        if session_date and info["start_date"] and info["end_date"]:
            if not info["start_date"] <= session_date <= info["end_date"]:
                errors.append(f"{prefix}: date is outside the event date range.")

        capacity_value = values.get("capacity")
        capacity = None
        if capacity_value not in (None, ""):
            try:
                capacity = int(capacity_value)
                if capacity < 0:
                    raise ValueError
            except (TypeError, ValueError):
                errors.append(f"{prefix}: capacity must be a positive whole number.")

        sessions.append(
            {
                "session_number": _string(values.get("session_number")),
                "session_name": session_name,
                "long_description": _string(values.get("long_description")),
                "resources_url": _url(
                    values.get("resources_url"), f"{prefix}: resources_url", errors
                ),
                "presenters": _string(values.get("presenters")),
                "date": session_date,
                "start_time": start_time,
                "end_time": end_time,
                "room": _string(values.get("room")),
                "capacity": capacity,
                "track": _string(values.get("track")),
                "session_type": _string(values.get("session_type")),
                "cech": _string(values.get("cech")),
                "featured": _boolean(values.get("featured"), default=False),
                "visible": _boolean(values.get("visible"), default=True),
            }
        )

    if not sessions:
        errors.append("Sessions: add at least one session.")

    sessions.sort(key=lambda item: (item["date"], item["start_time"], item["session_name"]))
    if errors:
        raise WorkbookValidationError(errors)

    return {"conference": info, "sessions": sessions}

